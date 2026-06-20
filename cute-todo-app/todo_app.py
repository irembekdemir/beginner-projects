import os
import customtkinter as ctk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ctk.set_appearance_mode("Light")  
ctk.set_default_color_theme("blue") 

PASTEL_PINK = "#F3E5F5"      # Background
NAVY = "#0D00FF"    # Buttons
TEXT_COLOR = "#FF0099"       # Font color

class TodoApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("MY TODO LIST ✨")
        self.geometry("450x550")
        self.configure(fg_color=PASTEL_PINK)
        self.resizable(False, False)

        self.title_label = ctk.CTkLabel(
            self, text="🌸 Daily Plan 🌸", 
            font=("Segoe UI", 22, "bold"), text_color=TEXT_COLOR
        )
        self.title_label.pack(pady=20)

        self.entry_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.entry_frame.pack(pady=10, padx=20, fill="x")

        self.task_entry = ctk.CTkEntry(
            self.entry_frame, placeholder_text="~Add new plan~", 
            width=260, font=("Segoe UI", 12), fg_color="#FFFFFF",
            text_color=TEXT_COLOR, border_color=NAVY
        )
        self.task_entry.pack(side="left", padx=(0, 10))
        self.task_entry.bind("<Return>", lambda event: self.add_task())

        self.add_button = ctk.CTkButton(
            self.entry_frame, text="Add ✨", width=80,
            fg_color=NAVY, hover_color="#B39DDB",
            text_color="#FFFFFF", font=("Segoe UI", 12, "bold"),
            command=self.add_task
        )
        self.add_button.pack(side="right")


        self.scrollable_frame = ctk.CTkScrollableFrame(
            self, width=380, height=300, 
            fg_color="#FFFFFF", border_color=NAVY, border_width=1
        )
        self.scrollable_frame.pack(pady=20, padx=20, fill="both", expand=True)


        self.button_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.button_frame.pack(pady=(0, 20), padx=20, fill="x")

        self.delete_button = ctk.CTkButton(
            self.button_frame, text="Delete the selected area! 🗑️", 
            fg_color="#FF0000", hover_color="#FF0000",
            text_color="#FFFFFF", font=("Segoe UI", 12, "bold"),
            command=self.delete_task
        )
        self.delete_button.pack(side="left", expand=True, fill="x", padx=(0, 5))

        self.export_button = ctk.CTkButton(
            self.button_frame, text="Save as an Excel File 📄", 
            fg_color="#00FF0D", hover_color="#66BB6A",
            text_color="#FFFFFF", font=("Segoe UI", 12, "bold"),
            command=self.export_to_excel
        )
        self.export_button.pack(side="right", expand=True, fill="x", padx=(5, 0))

        self.tasks = []

    def add_task(self):
        task_text = self.task_entry.get().strip()
        if task_text:
       
            cb = ctk.CTkCheckBox(
                self.scrollable_frame, text=task_text, 
                font=("Segoe UI", 13), text_color=TEXT_COLOR,
                checkmark_color="#FFFFFF", fg_color=NAVY, hover_color="#465AF1"
            )
            cb.pack(anchor="w", pady=10, padx=10)
            self.tasks.append(cb)
            self.task_entry.delete(0, 'end')
        else:
            messagebox.showwarning("Attention", "No empty area for a duty! 💕")

    def delete_task(self):
        for cb in self.tasks:
            if cb.get() == 1:
                cb.destroy()

        self.tasks = [
            cb for cb in self.tasks
            if cb.winfo_exists()
        ]

    def export_to_excel(self):
        if not self.tasks:
            messagebox.showwarning("Attention!", "Nothing to save here! 🌸")
            return

        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        excel_file = os.path.join(desktop_path, "todo_list.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "My Plansss"

        font_family = "Segoe UI"
        header_font = Font(name=font_family, size=12, bold=True, color="FFFFFF")
        data_font = Font(name=font_family, size=11, color="333333")
        
        header_fill = PatternFill(start_color="80DEEA", end_color="80DEEA", fill_type="solid") # Pastel blue/purple
        row_fill_1 = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Beige
        row_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        thin_border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )

        ws.append(["No", "Duty", "Status"])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border

        for idx, cb in enumerate(self.tasks, start=1):
            task_text = cb.cget("text")
            status = "Done! ✨" if cb.get() == 1 else "Will be done... ⏳"
            
            row_data = [idx, task_text, status]
            ws.append(row_data)
            
            current_row = ws.max_row
            row_fill = row_fill_1 if current_row % 2 == 0 else row_fill_2
            
            for col_idx, cell in enumerate(ws[current_row], start=1):
                cell.font = data_font
                cell.fill = row_fill
                cell.border = thin_border
                if col_idx == 2:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_center

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.row_dimensions[1].height = 25 

        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 20

        try:
            wb.save(excel_file)
            messagebox.showinfo("SUCCESS! 🎉", f"TODO LIST SAVED TO THE DESKTOP SUCCESSFULLY!\nFile: todo_list.xlsx")
        except Exception as e:
            messagebox.showerror("!ERROR!", f"An error occured during saving the file:\n{e}")

if __name__ == "__main__":
    app = TodoApp()
    app.mainloop()


